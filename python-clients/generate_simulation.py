#!/usr/bin/env python3
"""
generate_simulation.py

Generates an Excel workbook that simulates Phase 4 of the AI Round Table
Secret Sharing protocol — noise generation and masked submission — assuming
all partners have already completed Phases 1–3 (key registration, encapsulation,
decapsulation) and hold their pairwise shared secrets.

Usage:
    pip install openpyxl
    python generate_simulation.py [output.xlsx]

Output defaults to noise_simulation.xlsx in the current directory.
"""

import hashlib
import hmac as hmac_module
import struct
import sys
from itertools import combinations
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)

# ── Simulation parameters ──────────────────────────────────────────────────────

MAX_NOISE = 100_000_000

PARTNERS = [
    {"id": "partnerA", "name": "Partner A"},
    {"id": "partnerB", "name": "Partner B"},
    {"id": "partnerC", "name": "Partner C"},
]

# Example actual (private) MAU values per partner × (country, month).
# Partners never share these with each other or the aggregator.
SUBMISSIONS = [
    {
        "country": "USA", "month": "2026-05",
        "values": {"partnerA": 1_234_567, "partnerB": 987_654, "partnerC": 456_789},
    },
    {
        "country": "GBR", "month": "2026-05",
        "values": {"partnerA": 345_678, "partnerB": 234_567, "partnerC": 123_456},
    },
    {
        "country": "DEU", "month": "2026-05",
        "values": {"partnerA": 278_901, "partnerB": 198_234, "partnerC": 87_654},
    },
]

# ── Colour palette ─────────────────────────────────────────────────────────────

C_HEADER_DARK  = "1F3864"   # dark navy — section headers
C_HEADER_MID   = "2E75B6"   # mid blue — column headers
C_PARTNER_A    = "D6E4F0"   # light blue
C_PARTNER_B    = "D6EED6"   # light green
C_PARTNER_C    = "FFF3CD"   # light amber
C_PAIR_AB      = "E8D5F5"   # soft purple
C_PAIR_AC      = "D5EFF5"   # soft teal
C_PAIR_BC      = "FFE5CC"   # soft orange
C_TOTAL        = "E2EFDA"   # soft green — totals
C_WARNING      = "FF0000"   # red text for "aggregator sees this"
C_GOLD         = "FFD700"
C_WHITE        = "FFFFFF"
C_LIGHT_GREY   = "F2F2F2"

PARTNER_COLORS = {
    "partnerA": C_PARTNER_A,
    "partnerB": C_PARTNER_B,
    "partnerC": C_PARTNER_C,
}

PAIR_COLORS = {}

# ── Crypto helpers (identical to C# SecureNoiseGenerator + Python submit.py) ───

def simulated_shared_secret(id_a: str, id_b: str) -> bytes:
    """Deterministic simulated 32-byte shared secret for a partner pair."""
    smaller, larger = sorted([id_a, id_b])
    seed = f"simulation_seed|{smaller}|{larger}".encode()
    return hashlib.sha256(seed).digest()


def compute_noise(shared_secret: bytes, country: str, month_str: str) -> int:
    """
    Matches C# SecureNoiseGenerator.GenerateNoise and Python submit.py compute_noise.
    HMAC-SHA256(key=shared_secret, msg=f"{country}|{month_str}")
    → signed little-endian int64 from first 8 bytes
    → (seed % (2*MAX_NOISE+1)) - MAX_NOISE   [Python-style modulo, always ≥ 0]
    """
    msg = f"{country}|{month_str}".encode()
    h = hmac_module.new(shared_secret, msg, hashlib.sha256).digest()
    seed = struct.unpack("<q", h[:8])[0]
    noise_range = 2 * MAX_NOISE + 1
    return (seed % noise_range) - MAX_NOISE


def noise_sign(my_id: str, other_id: str) -> int:
    """Alphabetically smaller partner adds (+1), larger subtracts (−1)."""
    return 1 if my_id < other_id else -1


# ── Build simulation data ──────────────────────────────────────────────────────

def build_data():
    partner_ids = [p["id"] for p in PARTNERS]
    pairs = list(combinations(sorted(partner_ids), 2))  # (smaller, larger)

    # Shared secrets per pair
    secrets = {(a, b): simulated_shared_secret(a, b) for a, b in pairs}

    # Noise per pair per submission context
    rows = []
    for sub in SUBMISSIONS:
        country, month = sub["country"], sub["month"]
        pair_noise = {}
        for a, b in pairs:
            noise = compute_noise(secrets[(a, b)], country, month)
            pair_noise[(a, b)] = noise  # positive = A adds, B subtracts

        for partner in partner_ids:
            actual = sub["values"][partner]
            total_noise = 0
            noise_detail = {}
            for a, b in pairs:
                if partner not in (a, b):
                    continue
                n = pair_noise[(a, b)]
                sign = noise_sign(partner, b if partner == a else a)
                applied = n * sign
                other = b if partner == a else a
                noise_detail[other] = applied
                total_noise += applied

            rows.append({
                "country": country,
                "month": month,
                "partner": partner,
                "actual": actual,
                "noise_detail": noise_detail,   # {other_partner_id: applied_noise}
                "total_noise": total_noise,
                "masked": actual + total_noise,
            })

    return pairs, secrets, rows


# ── Excel style helpers ────────────────────────────────────────────────────────

def _side(style="thin"):
    return Side(style=style)


def thin_border():
    s = _side("thin")
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border():
    s = _side("medium")
    return Border(left=s, right=s, top=s, bottom=s)


def bottom_border(style="medium"):
    return Border(bottom=_side(style))


def set_cell(ws, row, col, value=None, *,
             bg=None, bold=False, fg="000000", size=11,
             align="left", wrap=False, num_fmt=None, bdr=None, italic=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic)
    h_align = {"left": "left", "center": "center", "right": "right"}.get(align, "left")
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    if bdr:
        cell.border = bdr
    return cell


def merge_header(ws, row, col_start, col_end, text, bg, fg="FFFFFF", size=12):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = set_cell(ws, row, col_start, text,
                    bg=bg, bold=True, fg=fg, size=size, align="center")
    cell.border = medium_border()
    return cell


def num(n):
    return f"{n:,}"


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    r = 1
    merge_header(ws, r, 1, 2,
                 "AI Round Table — Secure Aggregation Protocol  |  Phase 4 Simulation",
                 C_HEADER_DARK, size=14)
    ws.row_dimensions[r].height = 32
    r += 1

    sections = [
        ("What this workbook shows",
         "A step-by-step simulation of Phase 4: noise generation and masked submission. "
         "Phases 1–3 (key generation, encapsulation, decapsulation) are assumed complete — "
         "every partner already holds their pairwise shared secrets."),
        ("", ""),
        ("Protocol guarantee",
         "The aggregator learns ONLY the total across all partners. "
         "Individual partner values are cryptographically hidden — even with the "
         "masked submissions in hand, the aggregator cannot reverse-engineer any "
         "individual value without the ML-KEM shared secrets."),
        ("", ""),
        ("Why does noise magnitude not determine security?",
         "Security comes from the shared secret, not from noise size. "
         "The aggregator has no idea how much noise each partner applied — it could be "
         "+90 M or −3 M; without the shared secret it cannot tell. "
         "Noise magnitude matters only for protection against statistical inference "
         "across many repeated submissions over time. "
         "Increasing MAX_NOISE (currently 100,000,000) widens that window."),
        ("", ""),
        ("Sheets in this workbook", ""),
        ("  1. Overview",         "This page — protocol context and explanation."),
        ("  2. Shared Secrets",   "Simulated pairwise secrets (first 8 bytes shown as hex fingerprint)."),
        ("  3. Noise Calculation","HMAC-derived noise for every (partner-pair × country × month) combination."),
        ("  4. Masked Submissions","What each partner actually posts to the aggregator."),
        ("  5. Aggregation Proof","How the aggregator sums masked values and noise cancels perfectly."),
        ("", ""),
        ("MAX_NOISE",             f"{MAX_NOISE:,}  (±{MAX_NOISE/1e6:.0f} M — adjustable per deployment)"),
        ("Noise formula",
         "HMAC-SHA256(key=shared_secret, msg=country|YYYY-MM)\n"
         "→ signed little-endian int64 from first 8 bytes\n"
         "→ Python-style modulo → value in [−MAX_NOISE, +MAX_NOISE]"),
        ("Sign convention",
         "Alphabetically smaller partner ADDS the noise; larger partner SUBTRACTS it. "
         "Net contribution to the sum = 0 for every pair."),
    ]

    for label, desc in sections:
        ws.row_dimensions[r].height = 15 if desc else 8
        if label and desc:
            set_cell(ws, r, 1, label, bold=True, bg=C_LIGHT_GREY, bdr=thin_border())
            cell = set_cell(ws, r, 2, desc, bg=C_WHITE, bdr=thin_border(), wrap=True)
            ws.row_dimensions[r].height = max(15, desc.count("\n") * 15 + 15)
        elif label and not desc:
            merge_header(ws, r, 1, 2, label, C_HEADER_MID, size=11)
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


def build_secrets(wb, pairs, secrets):
    ws = wb.create_sheet("Shared Secrets")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 52

    r = 1
    merge_header(ws, r, 1, 5,
                 "Pairwise Shared Secrets  (simulated — in production these are derived via ML-KEM-768)",
                 C_HEADER_DARK, size=12)
    ws.row_dimensions[r].height = 28
    r += 1

    for col, label in enumerate(
        ["Smaller Partner", "Larger Partner", "Secret Fingerprint (first 8 B hex)",
         "Secret Length", "Security Note"], start=1
    ):
        set_cell(ws, r, col, label, bg=C_HEADER_MID, bold=True, fg="FFFFFF",
                 align="center", bdr=thin_border())
    r += 1

    pair_color_list = [C_PAIR_AB, C_PAIR_AC, C_PAIR_BC]
    for i, (a, b) in enumerate(pairs):
        color = pair_color_list[i % len(pair_color_list)]
        PAIR_COLORS[(a, b)] = color
        secret = secrets[(a, b)]
        fingerprint = secret[:8].hex().upper()
        formatted = " ".join(fingerprint[j:j+2] for j in range(0, len(fingerprint), 2))

        for col, (val, fmt) in enumerate([
            (a, {}),
            (b, {}),
            (formatted, {"align": "center"}),
            (f"{len(secret)} bytes (256 bits)", {"align": "center"}),
            ("Derived via ML-KEM-768 decapsulation. Never transmitted to aggregator.", {"italic": True}),
        ], start=1):
            set_cell(ws, r, col, val, bg=color, bdr=thin_border(), **fmt)
        r += 1

    r += 1
    merge_header(ws, r, 1, 5,
                 "Note: In production the full 32-byte secret is used as the HMAC key. "
                 "Only the fingerprint is shown here for readability.",
                 C_LIGHT_GREY, fg="444444", size=10)

    ws.sheet_view.showGridLines = False
    return ws


def build_noise(wb, pairs, secrets):
    ws = wb.create_sheet("Noise Calculation")

    pair_ids = [f"{a}↔{b}" for a, b in pairs]
    pair_color_list = [C_PAIR_AB, C_PAIR_AC, C_PAIR_BC]

    # Column layout: Country | Month | [per-pair: raw noise, sign_A, sign_B] ...
    # We'll show: for each pair — raw noise value; then applied noise for each partner
    col_country = 1
    col_month   = 2
    pair_col_start = 3   # each pair takes 3 cols: raw noise, applied to smaller, applied to larger

    total_cols = 2 + len(pairs) * 3

    ws.column_dimensions[get_column_letter(col_country)].width = 10
    ws.column_dimensions[get_column_letter(col_month)].width = 12
    for i, (a, b) in enumerate(pairs):
        base = pair_col_start + i * 3
        ws.column_dimensions[get_column_letter(base)].width = 22
        ws.column_dimensions[get_column_letter(base + 1)].width = 22
        ws.column_dimensions[get_column_letter(base + 2)].width = 22

    r = 1
    merge_header(ws, r, 1, total_cols,
                 "Noise Calculation — HMAC-SHA256(shared_secret, country|month) for every pair × submission context",
                 C_HEADER_DARK, size=12)
    ws.row_dimensions[r].height = 28
    r += 1

    # Pair group headers
    set_cell(ws, r, col_country, "", bg=C_HEADER_MID)
    set_cell(ws, r, col_month,   "", bg=C_HEADER_MID)
    for i, (a, b) in enumerate(pairs):
        base = pair_col_start + i * 3
        color = pair_color_list[i % len(pair_color_list)]
        merge_header(ws, r, base, base + 2,
                     f"Pair  {a} ↔ {b}", color, fg=C_HEADER_DARK, size=10)
    r += 1

    # Column sub-headers
    set_cell(ws, r, col_country, "Country", bg=C_HEADER_MID, bold=True, fg="FFFFFF",
             align="center", bdr=thin_border())
    set_cell(ws, r, col_month,   "Month",   bg=C_HEADER_MID, bold=True, fg="FFFFFF",
             align="center", bdr=thin_border())
    for i, (a, b) in enumerate(pairs):
        base = pair_col_start + i * 3
        color = pair_color_list[i % len(pair_color_list)]
        set_cell(ws, r, base,     "Raw Noise (shared)",
                 bg=color, bold=True, align="center", bdr=thin_border(), wrap=True)
        set_cell(ws, r, base + 1, f"{a} applies\n(+, smaller)",
                 bg=PARTNER_COLORS["partnerA"], bold=True, align="center", bdr=thin_border(), wrap=True)
        set_cell(ws, r, base + 2, f"{b} applies\n(−, larger)",
                 bg=PARTNER_COLORS.get(b, C_LIGHT_GREY), bold=True, align="center", bdr=thin_border(), wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    prev_context = None
    for sub in SUBMISSIONS:
        country, month = sub["country"], sub["month"]
        context = (country, month)
        row_bg = C_WHITE if prev_context != context else C_LIGHT_GREY
        prev_context = context

        set_cell(ws, r, col_country, country, bg=row_bg, align="center", bdr=thin_border())
        set_cell(ws, r, col_month,   month,   bg=row_bg, align="center", bdr=thin_border())

        for i, (a, b) in enumerate(pairs):
            base = pair_col_start + i * 3
            color = pair_color_list[i % len(pair_color_list)]
            noise = compute_noise(secrets[(a, b)], country, month)
            applied_a = noise  * noise_sign(a, b)   # smaller adds
            applied_b = noise  * noise_sign(b, a)   # larger subtracts

            fmt = "#,##0;[Red]-#,##0"
            set_cell(ws, r, base,     noise,     bg=color, align="right", num_fmt=fmt, bdr=thin_border())
            set_cell(ws, r, base + 1, applied_a, bg=PARTNER_COLORS["partnerA"],
                     align="right", num_fmt=fmt, bdr=thin_border())
            set_cell(ws, r, base + 2, applied_b,
                     bg=PARTNER_COLORS.get(b, C_LIGHT_GREY),
                     align="right", num_fmt=fmt, bdr=thin_border())
        r += 1

    r += 1
    merge_header(ws, r, 1, total_cols,
                 "Raw Noise is the same value for both partners in a pair — "
                 "the smaller partner adds it, the larger subtracts it. Net effect on the aggregate = 0.",
                 C_LIGHT_GREY, fg="444444", size=10)

    ws.sheet_view.showGridLines = False
    return ws


def build_submissions(wb, rows):
    ws = wb.create_sheet("Masked Submissions")

    partner_ids = [p["id"] for p in PARTNERS]
    other_ids_per_partner = {
        pid: [o for o in partner_ids if o != pid]
        for pid in partner_ids
    }

    # Columns: Country | Month | Partner | Actual Value (SECRET) |
    #          noise with B | noise with C (or A/C for B, etc.) |
    #          Total Noise | Masked Value (POSTED TO AGGREGATOR)
    max_others = max(len(v) for v in other_ids_per_partner.values())

    col_country = 1
    col_month   = 2
    col_partner = 3
    col_actual  = 4
    noise_cols_start = 5
    col_total_noise = noise_cols_start + max_others
    col_masked  = col_total_noise + 1
    total_cols  = col_masked

    widths = [10, 12, 14, 22] + [22] * max_others + [16, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    merge_header(ws, r, 1, total_cols,
                 "Masked Submissions — What Each Partner Posts to the Aggregator",
                 C_HEADER_DARK, size=12)
    ws.row_dimensions[r].height = 28
    r += 1

    headers = ["Country", "Month", "Partner",
               "Actual MAU\n(SECRET — never shared)",
               "Noise with Partner 2", "Noise with Partner 3",
               "Total Noise", "Masked MAU\n(POSTED TO AGGREGATOR)"]
    for col, label in enumerate(headers, start=1):
        set_cell(ws, r, col, label, bg=C_HEADER_MID, bold=True, fg="FFFFFF",
                 align="center", bdr=thin_border(), wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    num_fmt = "#,##0;[Red]-#,##0"
    prev_context = None

    for row in rows:
        context = (row["country"], row["month"])
        pid = row["partner"]
        color = PARTNER_COLORS.get(pid, C_WHITE)
        alt_bg = C_LIGHT_GREY if prev_context == context else C_WHITE

        set_cell(ws, r, col_country, row["country"], bg=alt_bg,  align="center", bdr=thin_border())
        set_cell(ws, r, col_month,   row["month"],   bg=alt_bg,  align="center", bdr=thin_border())
        set_cell(ws, r, col_partner, pid,             bg=color,   align="center", bold=True, bdr=thin_border())
        set_cell(ws, r, col_actual,  row["actual"],   bg=color,   align="right",  num_fmt=num_fmt, bdr=thin_border())

        others = sorted(row["noise_detail"].keys())
        for j, other in enumerate(others):
            applied = row["noise_detail"][other]
            set_cell(ws, r, noise_cols_start + j, applied,
                     bg=color, align="right", num_fmt=num_fmt, bdr=thin_border())
        for j in range(len(others), max_others):
            set_cell(ws, r, noise_cols_start + j, "—", bg=color, align="center", bdr=thin_border())

        set_cell(ws, r, col_total_noise, row["total_noise"],
                 bg=color, align="right", num_fmt=num_fmt, bdr=thin_border(), bold=True)
        set_cell(ws, r, col_masked, row["masked"],
                 bg="FFD700", align="right", num_fmt=num_fmt, bdr=medium_border(), bold=True)

        prev_context = context
        r += 1

    r += 1
    merge_header(ws, r, 1, total_cols,
                 "The aggregator receives ONLY the gold 'Masked MAU' column. "
                 "It cannot see actual values, individual noise contributions, or total noise.",
                 "FFF9E6", fg="7D6608", size=10)

    ws.sheet_view.showGridLines = False
    return ws


def build_aggregation(wb, rows):
    ws = wb.create_sheet("Aggregation Proof")

    partner_ids = [p["id"] for p in PARTNERS]
    contexts = list({(r["country"], r["month"]) for r in rows})
    contexts.sort()

    # Columns: Country | Month | [per partner: actual | masked] | Sum Actual | Sum Masked | Diff (=0) | Proof
    n = len(partner_ids)
    col_country     = 1
    col_month       = 2
    partner_cols    = {}  # pid -> (col_actual, col_masked)
    base = 3
    for pid in partner_ids:
        partner_cols[pid] = (base, base + 1)
        base += 2
    col_sum_actual  = base
    col_sum_masked  = base + 1
    col_diff        = base + 2
    col_proof       = base + 3
    total_cols      = col_proof

    ws.column_dimensions[get_column_letter(col_country)].width = 10
    ws.column_dimensions[get_column_letter(col_month)].width   = 12
    for pid in partner_ids:
        ca, cm = partner_cols[pid]
        ws.column_dimensions[get_column_letter(ca)].width = 18
        ws.column_dimensions[get_column_letter(cm)].width = 18
    ws.column_dimensions[get_column_letter(col_sum_actual)].width  = 18
    ws.column_dimensions[get_column_letter(col_sum_masked)].width  = 18
    ws.column_dimensions[get_column_letter(col_diff)].width        = 14
    ws.column_dimensions[get_column_letter(col_proof)].width       = 20

    r = 1
    merge_header(ws, r, 1, total_cols,
                 "Aggregation Proof — Noise Cancels, True Total Revealed",
                 C_HEADER_DARK, size=12)
    ws.row_dimensions[r].height = 28
    r += 1

    # Partner group headers
    set_cell(ws, r, col_country, "", bg=C_HEADER_MID)
    set_cell(ws, r, col_month,   "", bg=C_HEADER_MID)
    for pid in partner_ids:
        ca, cm = partner_cols[pid]
        merge_header(ws, r, ca, cm, pid, PARTNER_COLORS.get(pid, C_WHITE), fg=C_HEADER_DARK, size=10)
    for col, label in [
        (col_sum_actual, "True Total\n(unknown to aggregator)"),
        (col_sum_masked, "Aggregated Total\n(aggregator computes this)"),
        (col_diff,       "Difference\n(must = 0)"),
        (col_proof,      "Result"),
    ]:
        set_cell(ws, r, col, label, bg=C_TOTAL, bold=True, align="center",
                 bdr=thin_border(), wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    # Sub-headers
    set_cell(ws, r, col_country, "Country", bg=C_HEADER_MID, bold=True, fg="FFFFFF",
             align="center", bdr=thin_border())
    set_cell(ws, r, col_month,   "Month",   bg=C_HEADER_MID, bold=True, fg="FFFFFF",
             align="center", bdr=thin_border())
    for pid in partner_ids:
        ca, cm = partner_cols[pid]
        color = PARTNER_COLORS.get(pid, C_WHITE)
        set_cell(ws, r, ca, "Actual (secret)", bg=color, bold=True,
                 align="center", bdr=thin_border(), wrap=True)
        set_cell(ws, r, cm, "Masked (posted)",  bg="FFD700", bold=True,
                 align="center", bdr=thin_border(), wrap=True)
    for col in [col_sum_actual, col_sum_masked, col_diff, col_proof]:
        set_cell(ws, r, col, "", bg=C_TOTAL, bdr=thin_border())
    ws.row_dimensions[r].height = 28
    r += 1

    num_fmt = "#,##0"

    for country, month in contexts:
        ctx_rows = {row["partner"]: row for row in rows
                    if row["country"] == country and row["month"] == month}

        sum_actual = sum(ctx_rows[pid]["actual"] for pid in partner_ids)
        sum_masked = sum(ctx_rows[pid]["masked"] for pid in partner_ids)
        diff = sum_masked - sum_actual

        set_cell(ws, r, col_country, country, align="center", bdr=thin_border())
        set_cell(ws, r, col_month,   month,   align="center", bdr=thin_border())

        for pid in partner_ids:
            ca, cm = partner_cols[pid]
            color = PARTNER_COLORS.get(pid, C_WHITE)
            set_cell(ws, r, ca, ctx_rows[pid]["actual"], bg=color,
                     align="right", num_fmt=num_fmt, bdr=thin_border())
            set_cell(ws, r, cm, ctx_rows[pid]["masked"], bg="FFD700",
                     align="right", num_fmt=num_fmt, bdr=thin_border(), bold=True)

        set_cell(ws, r, col_sum_actual, sum_actual, bg=C_TOTAL,
                 align="right", num_fmt=num_fmt, bdr=thin_border(), bold=True)
        set_cell(ws, r, col_sum_masked, sum_masked, bg=C_TOTAL,
                 align="right", num_fmt=num_fmt, bdr=thin_border(), bold=True)
        set_cell(ws, r, col_diff, diff, bg=C_TOTAL,
                 align="right", num_fmt=num_fmt, bdr=medium_border(), bold=True,
                 fg="006400" if diff == 0 else "FF0000")
        set_cell(ws, r, col_proof,
                 "✓  Noise cancelled!" if diff == 0 else "✗  ERROR",
                 bg=C_TOTAL if diff == 0 else "FFE0E0",
                 bold=True, align="center", bdr=medium_border(),
                 fg="006400" if diff == 0 else "FF0000")
        r += 1

    r += 2
    merge_header(ws, r, 1, total_cols,
                 "Mathematical proof: for every pair (A, B) with A < B, "
                 "A adds +noise and B adds −noise. Summing all partners: Σ noise = 0.  "
                 "Therefore: Σ masked_values = Σ actual_values. ✓",
                 C_TOTAL, fg="006400", size=11)

    ws.sheet_view.showGridLines = False
    return ws


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    out_path = Path(sys.argv[1] if len(sys.argv) > 1 else "noise_simulation.xlsx")

    print("Building simulation data...")
    pairs, secrets, rows = build_data()

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    print("Building sheet: Overview")
    build_overview(wb)

    print("Building sheet: Shared Secrets")
    build_secrets(wb, pairs, secrets)

    print("Building sheet: Noise Calculation")
    build_noise(wb, pairs, secrets)

    print("Building sheet: Masked Submissions")
    build_submissions(wb, rows)

    print("Building sheet: Aggregation Proof")
    build_aggregation(wb, rows)

    wb.save(out_path)
    print(f"\nSaved → {out_path.resolve()}")
    print("Open in Excel or LibreOffice Calc.")


if __name__ == "__main__":
    main()
